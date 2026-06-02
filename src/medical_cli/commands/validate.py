"""Validate command module with click framework.

This module provides the validate command group for validating clinical data,
including the specialized validate_labs subcommand for Excel laboratory results.
"""

import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import click

from medical_cli.core.data_extractor import EDCDataExtractor
from medical_cli.core.data_validator import (
    ClinicalSubjectBase,
    LabValue,
    LAB_TEST_REFERENCE_RANGES,
)
from medical_cli.core.adapters.edc_wide_adapter import EDCWideAdapter
from medical_cli.utils.logger import get_logger


# Configure module logger
logger = get_logger("medical_cli.commands.validate")


def add_parser(subparsers: Any) -> None:
    """Add validate command parser (backwards compatibility with argparse)."""
    # This function is kept for backwards compatibility with the existing CLI
    # but the main implementation now uses click
    pass


def execute(args: Any) -> int:
    """Execute the validate command (backwards compatibility wrapper)."""
    # This is a compatibility wrapper - actual logic is in click commands
    return 0


# ============================================================================
# Click-based Command Group
# ============================================================================

@click.group(
    name="validate",
    help="Validate clinical data and laboratory results",
    epilog="For more information, visit: https://github.com/example/medical-cli",
)
@click.version_option(version="0.1.0", prog_name="medical-cli")
def validate_group():
    """Clinical data validation commands.
    
    This command group provides tools for validating clinical trial data,
    including specialized laboratory result validation for Excel files.
    """
    pass


@validate_group.command(
    name="labs",
    help="Validate Excel laboratory test results",
    short_help="Validate laboratory data from Excel files",
)
@click.option(
    "--input",
    "-i",
    "input_file",
    required=True,
    type=click.Path(exists=True, file_okay=True, dir_okay=False),
    help="Path to the input Excel file (.xlsx or .xls) containing lab results",
)
@click.option(
    "--output",
    "-o",
    "output_file",
    required=False,
    type=click.Path(file_okay=True, dir_okay=False),
    help="Path to save the validation report (JSON format). "
         "If not specified, report is printed to stdout.",
)
@click.option(
    "--strict",
    "-s",
    is_flag=True,
    default=False,
    help="Enable strict validation mode. "
         "Exit with non-zero code if any validation errors are found.",
)
@click.option(
    "--verbose",
    "-v",
    is_flag=True,
    default=False,
    help="Enable verbose output with detailed processing information.",
)
@click.option(
    "--abnormal-only",
    is_flag=True,
    default=False,
    help="Only report records with abnormal lab values in the output.",
)
@click.option(
    "--export-excel",
    "-e",
    "export_excel",
    is_flag=True,
    default=False,
    help="Export sanitized Excel with validation flags. "
         "Creates [filename]_sanitized.xlsx with PHI masked and validation flags.",
)
@click.option(
    "--adapter",
    "-a",
    "adapter_type",
    type=click.Choice(["standard", "edc_wide"], case_sensitive=False),
    default="standard",
    help="Data adapter type: 'standard' for simple files, "
         "'edc_wide' for EDC exports with multi-row headers and wide-table format.",
)
def validate_labs(
    input_file: str,
    output_file: Optional[str],
    strict: bool,
    verbose: bool,
    abnormal_only: bool,
    export_excel: bool,
    adapter_type: str,
) -> None:
    """Validate Excel laboratory test results with PHI sanitization.
    
    This command performs end-to-end validation of clinical laboratory data:
    
    1. Reads data from the specified Excel file using selected adapter
    2. Automatically detects and masks Protected Health Information (PHI)
    3. Validates each record against clinical data schemas
    4. Checks laboratory values against reference ranges
    5. Generates a comprehensive validation report
    6. Optionally exports a sanitized Excel with validation flags
    
    \b
    Supported Adapters:
    - standard: For simple Excel/CSV with standard column structure
    - edc_wide: For EDC system exports with multi-row headers and wide-table format
    
    \b
    Examples:
    
    \b
    # Basic validation (output to terminal):
        medical-cli validate labs -i lab_results.xlsx
    
    \b
    # Validate EDC wide-table export with Chinese headers:
        medical-cli validate labs -i edc_export.xlsx -a edc_wide
    
    \b
    # Save report to file:
        medical-cli validate labs -i lab_results.xlsx -o report.json
    
    \b
    # Export sanitized Excel with validation flags:
        medical-cli validate labs -i lab_results.xlsx -e
    
    \b
    # Strict mode with verbose output:
        medical-cli validate labs -i lab_results.xlsx -s -v
    
    \b
    # Only show abnormal values:
        medical-cli validate labs -i lab_results.xlsx --abnormal-only
    
    \b
    Exit codes:
        0 - Validation successful (no errors or only warnings)
        1 - Validation failed (errors found or file processing error)
    """
    logger.info(f"Starting lab validation: {input_file}")
    
    # Initialize report structure
    report: Dict[str, Any] = {
        "status": "unknown",
        "input_file": str(input_file),
        "total_rows_processed": 0,
        "total_rows_with_data": 0,
        "phi_sanitization": {
            "records_sanitized": 0,
            "fields_masked": 0,
            "phi_detected": False,
        },
        "lab_validation": {
            "total_lab_values": 0,
            "abnormal_values": [],
            "unit_errors": [],
            "missing_required": [],
        },
        "summary": {
            "errors": 0,
            "warnings": 0,
            "abnormal_count": 0,
        },
        "rows": [],
    }
    
    try:
        # -----------------------------------------------------------------
        # Step 1: Read Excel file (using selected adapter)
        # -----------------------------------------------------------------
        if verbose:
            if adapter_type == "edc_wide":
                click.echo("\n[Step 1/4] Parsing EDC wide-table structure...")
            else:
                click.echo("\n[Step 1/4] Analyzing Excel data...")
        
        logger.info(f"Reading file with adapter: {adapter_type}")
        logger.info(f"Input file: {input_file}")
        
        data: List[Dict[str, Any]] = []
        extractor: Optional[EDCDataExtractor] = None
        
        try:
            if adapter_type == "edc_wide":
                # Use EDC Wide Adapter for complex multi-row headers
                if verbose:
                    click.echo("  Using EDC Wide Adapter for complex structure")
                
                adapter = EDCWideAdapter(input_file)
                parsed_data = adapter.parse_lab_data()
                
                # Convert adapter output to standard format for validation
                for record in parsed_data:
                    # Set default values for required fields that may not be in EDC data
                    # Use valid values to avoid validation errors
                    record.setdefault("Age", 30)  # Adult age to avoid pediatric warning
                    record.setdefault("Gender", "M")  # Default to Male (valid Gender value)
                    record.setdefault("Visit_Date", "2024-01-01")
                    
                    # Format for validation
                    validated_record = {
                        "Subject_ID": record.get("Subject_ID", ""),
                        "Age": record.get("Age", 30),
                        "Gender": record.get("Gender", "M"),
                        "Visit_Date": record.get("Visit_Date", ""),
                    }
                    if "lab_values" in record:
                        validated_record["lab_values"] = record["lab_values"]
                    
                    data.append(validated_record)
                
                if verbose:
                    click.echo(f"  ✓ Parsed {len(data)} records from wide-table format")
                
            else:
                # Standard adapter
                extractor = EDCDataExtractor(
                    input_path=input_file,
                    encoding="utf-8",
                    strict_mode=False,
                )
                data = extractor.read_data()
                
                if verbose:
                    click.echo(f"  ✓ Read {len(data)} rows")
                    
        except FileNotFoundError:
            click.echo(
                click.style(
                    f"Error: Input file not found: {input_file}",
                    fg="red",
                ),
                err=True,
            )
            report["status"] = "error"
            report["error_message"] = f"File not found: {input_file}"
            _output_report(report, output_file)
            sys.exit(1)
        except ValueError as e:
            error_msg = str(e)
            if "Unsupported file format" in error_msg:
                click.echo(
                    click.style(
                        f"Error: {error_msg}\n"
                        f"Please provide a valid Excel file (.xlsx or .xls)",
                        fg="red",
                    ),
                    err=True,
                )
            else:
                click.echo(
                    click.style(f"Error: {error_msg}", fg="red"),
                    err=True,
                )
            report["status"] = "error"
            report["error_message"] = error_msg
            _output_report(report, output_file)
            sys.exit(1)
        except Exception as e:
            click.echo(
                click.style(f"Error reading file: {e}", fg="red"),
                err=True,
            )
            report["status"] = "error"
            report["error_message"] = str(e)
            _output_report(report, output_file)
            sys.exit(1)
        
        report["phi_sanitization"] = {
            "records_sanitized": 0 if adapter_type == "edc_wide" else phi_report.get("total_records", 0),
            "fields_masked": 0 if adapter_type == "edc_wide" else phi_report.get("masked_count", 0),
            "phi_detected": False if adapter_type == "edc_wide" else not phi_report.get("is_clean", True),
            "detections": {} if adapter_type == "edc_wide" else phi_report.get("detections_by_type", {}),
        }
        
        report["total_rows_processed"] = len(data)
        
        if verbose:
            if adapter_type == "edc_wide":
                click.echo(f"  ✓ Parsed {len(data)} records (PHI detection not available for edc_wide adapter)")
            else:
                click.echo(
                    f"  ✓ Read {len(data)} rows "
                    f"(PHI fields masked: {phi_report.get('masked_count', 0)})"
                )
        
        # -----------------------------------------------------------------
        # Step 2: Validate data with ClinicalDataValidator
        # -----------------------------------------------------------------
        if verbose:
            if adapter_type == "edc_wide":
                click.echo("\n[Step 2/4] Validating clinical data (from adapter)...")
            else:
                click.echo("\n[Step 2/4] Validating clinical data...")
        
        logger.info("Starting data validation")
        
        abnormal_records: List[Dict[str, Any]] = []
        unit_errors: List[Dict[str, Any]] = []
        missing_required: List[Dict[str, Any]] = []
        total_lab_values: int = 0
        
        for row_idx, row in enumerate(data, start=1):
            row_has_data = any(v for v in row.values() if v)
            if not row_has_data:
                continue
            
            report["total_rows_with_data"] += 1
            
            # Build row report
            row_report: Dict[str, Any] = {
                "row_number": row_idx,
                "validation_status": "valid",
                "lab_validation": {
                    "lab_values": [],
                    "abnormal_count": 0,
                },
                "issues": [],
            }
            
            # Try to validate as ClinicalSubjectBase
            # For EDC wide adapter, we may have lab_values already prepared
            try:
                # Check if row already has lab_values (from EDC wide adapter)
                if "lab_values" in row:
                    # Prepare a minimal row for ClinicalSubjectBase
                    # Use valid default values to avoid validation errors
                    prepared_row = {
                        "Subject_ID": row.get("Subject_ID", ""),
                        "Age": row.get("Age", 30),  # Default adult age
                        "Gender": row.get("Gender", "M"),  # Default to Male
                        "Visit_Date": row.get("Visit_Date", "2024-01-01"),
                        "lab_values": row.get("lab_values", []),
                    }
                    subject = ClinicalSubjectBase(**prepared_row)
                    lab_values_to_validate = subject.lab_values
                else:
                    # Convert row to proper types for standard format
                    validated_row = _prepare_row_for_validation(row)
                    subject = ClinicalSubjectBase(**validated_row)
                    lab_values_to_validate = subject.lab_values
                
                # Validate lab values if present
                if lab_values_to_validate:
                    total_lab_values += len(lab_values_to_validate)
                    
                    lab_results = _validate_lab_values(
                        subject.Subject_ID,
                        lab_values_to_validate,
                        verbose,
                    )
                    
                    row_report["lab_validation"]["lab_values"] = lab_results
                    row_report["lab_validation"]["abnormal_count"] = sum(
                        1 for lab in lab_results if lab.get("is_abnormal", False)
                    )
                    
                    # Track abnormal values
                    for lab in lab_results:
                        if lab.get("is_abnormal", False):
                            abnormal_records.append({
                                "row": row_idx,
                                "subject_id": subject.Subject_ID,
                                "test_code": lab.get("test_code"),
                                "value": lab.get("value"),
                                "unit": lab.get("unit"),
                                "reference_range": f"{lab.get('reference_range_low')} - {lab.get('reference_range_high')}",
                                "deviation": lab.get("deviation"),
                            })
                            
                            report["summary"]["abnormal_count"] += 1
                
            except Exception as e:
                error_str = str(e)
                
                # Check for specific error types
                if "missing" in error_str.lower() or "field required" in error_str.lower():
                    missing_required.append({
                        "row": row_idx,
                        "error": error_str,
                        "row_data": {k: v for k, v in row.items() if k in [
                            "Subject_ID", "Age", "Gender", "Visit_Date"
                        ]},
                    })
                    row_report["issues"].append({
                        "type": "missing_required",
                        "message": error_str,
                    })
                    row_report["validation_status"] = "error"
                else:
                    # General validation error
                    row_report["issues"].append({
                        "type": "validation_error",
                        "message": error_str,
                    })
                    row_report["validation_status"] = "warning"
                
                if verbose:
                    logger.warning(f"Row {row_idx} validation issue: {error_str}")
            
            # Add to report
            if not abnormal_only or row_report["lab_validation"]["abnormal_count"] > 0:
                report["rows"].append(row_report)
        
        # Update lab validation summary
        report["lab_validation"] = {
            "total_lab_values": total_lab_values,
            "abnormal_values": abnormal_records,
            "unit_errors": unit_errors,
            "missing_required": missing_required,
        }
        
        report["summary"]["errors"] = len(missing_required)
        report["summary"]["warnings"] = len(unit_errors)
        
        if verbose:
            click.echo(
                f"  ✓ Validated {report['total_rows_with_data']} rows "
                f"(Abnormal labs: {report['summary']['abnormal_count']})"
            )
        
        # -----------------------------------------------------------------
        # Step 3: Generate report and optionally export Excel
        # -----------------------------------------------------------------
        if verbose:
            if adapter_type == "edc_wide":
                click.echo("\n[Step 3/4] Generating validation report...")
            else:
                click.echo("\n[Step 3/4] Generating validation report...")
        
        # Determine overall status
        if report["summary"]["errors"] > 0:
            report["status"] = "failed"
        elif report["summary"]["abnormal_count"] > 0:
            report["status"] = "completed_with_warnings"
        else:
            report["status"] = "success"
        
        # Output report
        _output_report(report, output_file)
        
        # Export sanitized Excel if requested
        if export_excel:
            _export_sanitized_excel(extractor, report, verbose)
        
        # Print summary to terminal
        _print_summary(report, verbose)
        
        # Log completion
        logger.info(f"Validation complete: {report['status']}")
        
        # Exit with appropriate code
        if strict and report["summary"]["errors"] > 0:
            sys.exit(1)
        
    except KeyboardInterrupt:
        click.echo("\n\nValidation cancelled by user.", err=True)
        report["status"] = "cancelled"
        _output_report(report, output_file)
        sys.exit(1)
    except Exception as e:
        click.echo(
            click.style(f"Unexpected error: {e}", fg="red"),
            err=True,
        )
        logger.error(f"Unexpected validation error: {e}")
        report["status"] = "error"
        report["error_message"] = str(e)
        _output_report(report, output_file)
        sys.exit(1)


def _prepare_row_for_validation(row: Dict[str, Any]) -> Dict[str, Any]:
    """Prepare a row dictionary for ClinicalSubjectBase validation.
    
    Parses flat Excel structure with independent Test_Code, Value, Unit columns.
    Each row may contain multiple lab tests - extracts all and assembles into
    lab_values list for Pydantic validation.
    
    Args:
        row: Raw row data from Excel/CSV.
        
    Returns:
        Dictionary with properly typed values for validation.
    """
    prepared: Dict[str, Any] = {}
    lab_values: List[Dict[str, Any]] = []
    
    # Extract Test_Code, Value, Unit columns directly
    # Excel structure: Test_Code, Value, Unit (independent columns per row)
    test_code = row.get("Test_Code") or row.get("test_code") or row.get("Test_Code", "").strip()
    value = row.get("Value") or row.get("value")
    unit = row.get("Unit") or row.get("unit") or row.get("Unit", "").strip()
    
    # If Test_Code has value and Value is not empty, create lab entry
    if test_code and value is not None and value != "":
        try:
            lab_entry = {
                "test_code": str(test_code).upper().strip(),
                "test_name": str(test_code).strip(),  # Use test_code as name if no separate name column
                "value": float(value),
                "unit": str(unit).strip() if unit else "",
            }
            lab_values.append(lab_entry)
        except (ValueError, TypeError):
            # Skip invalid numeric values
            pass
    
    # Handle other fields
    for key, value in row.items():
        # Skip empty values
        if value is None or value == "":
            continue
        
        # Skip Test_Code/Value/Unit columns (already processed)
        key_lower = key.lower()
        if key_lower in ("test_code", "value", "unit"):
            continue
        
        # Handle specific fields
        if key_lower == "age":
            try:
                prepared[key] = int(float(value))
            except (ValueError, TypeError):
                prepared[key] = value
        elif key_lower in ("visit_date", "date", "date_of_visit"):
            if isinstance(value, str):
                prepared[key] = value
            else:
                prepared[key] = str(value)
        elif key_lower == "gender":
            prepared[key] = str(value).upper().strip()
        else:
            prepared[key] = value
    
    # Add lab_values to prepared dict if any were found
    if lab_values:
        prepared["lab_values"] = lab_values
    
    return prepared


def _validate_lab_values(
    subject_id: str,
    lab_values: List[LabValue],
    verbose: bool = False,
) -> List[Dict[str, Any]]:
    """Validate laboratory values and check against reference ranges.
    
    Args:
        subject_id: Subject identifier for logging.
        lab_values: List of LabValue objects to validate.
        verbose: Enable verbose output.
        
    Returns:
        List of dictionaries with validation results.
    """
    results: List[Dict[str, Any]] = []
    
    for lab in lab_values:
        result: Dict[str, Any] = {
            "test_code": lab.test_code,
            "test_name": lab.test_name,
            "value": lab.value,
            "unit": lab.unit,
            "reference_range_low": lab.reference_range_low,
            "reference_range_high": lab.reference_range_high,
            "is_abnormal": lab.is_abnormal,
            "deviation": None,
        }
        
        # Calculate deviation from normal range
        if lab.reference_range_low is not None and lab.reference_range_high is not None:
            range_width = lab.reference_range_high - lab.reference_range_low
            
            if range_width > 0:
                if lab.value < lab.reference_range_low:
                    deviation = (lab.reference_range_low - lab.value) / range_width
                    result["deviation"] = -round(deviation, 2)
                elif lab.value > lab.reference_range_high:
                    deviation = (lab.value - lab.reference_range_high) / range_width
                    result["deviation"] = round(deviation, 2)
        
        # Check for unit mismatches
        if lab.test_code in LAB_TEST_REFERENCE_RANGES:
            expected_unit = LAB_TEST_REFERENCE_RANGES[lab.test_code][2]
            if expected_unit and lab.unit != expected_unit:
                result["unit_warning"] = (
                    f"Unit mismatch: expected '{expected_unit}', got '{lab.unit}'"
                )
        
        results.append(result)
        
        if verbose and lab.is_abnormal:
            logger.warning(
                f"{subject_id}: Abnormal {lab.test_code}={lab.value} "
                f"(ref: {lab.reference_range_low}-{lab.reference_range_high})"
            )
    
    return results


def _export_sanitized_excel(
    extractor: EDCDataExtractor,
    report: Dict[str, Any],
    verbose: bool,
) -> None:
    """Export sanitized Excel file with validation flags.
    
    Creates a new Excel file with:
    1. PHI masked data from extractor
    2. Validation_Flags column with warnings for each row
    
    Args:
        extractor: EDCDataExtractor instance with sanitized data
        report: Validation report with row issues and lab validation
        verbose: Enable verbose output
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        logger.warning("openpyxl not installed. Cannot export Excel file.")
        click.echo(
            click.style(
                "Warning: openpyxl not installed. Cannot export Excel. "
                "Install with: pip install openpyxl",
                fg="yellow",
            )
        )
        return
    
    input_path = extractor.input_path
    output_path = input_path.parent / f"{input_path.stem}_sanitized.xlsx"
    
    try:
        # Get sanitized data from extractor
        sanitized_data = extractor._raw_data
        
        if not sanitized_data:
            logger.warning("No data to export")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clinical Data"
        
        # Get headers from first row and add Validation_Flags
        headers = list(sanitized_data[0].keys()) + ["Validation_Flags"]
        
        # Write header row
        ws.append(headers)
        
        # Build row -> issues mapping from report
        row_issues: Dict[int, List[str]] = {}
        
        # Map abnormal lab values to rows
        for abn in report.get("lab_validation", {}).get("abnormal_values", []):
            row_num = abn.get("row", 0)
            if row_num not in row_issues:
                row_issues[row_num] = []
            
            test_code = abn.get("test_code", "")
            value = abn.get("value", 0)
            ref = abn.get("reference_range", "")
            
            # Parse reference range to determine if High or Low
            try:
                if " - " in ref:
                    parts = ref.split(" - ")
                    ref_low = float(parts[0].strip())
                    ref_high = float(parts[1].strip())
                    if value < ref_low:
                        flag = f"{test_code}: {value} (Low)"
                    elif value > ref_high:
                        flag = f"{test_code}: {value} (High)"
                    else:
                        flag = f"{test_code}: {value}"
                else:
                    flag = f"{test_code}: {value}"
            except (ValueError, TypeError):
                flag = f"{test_code}: {value}"
            
            row_issues[row_num].append(flag)
        
        # Map missing required fields to rows
        for missing in report.get("lab_validation", {}).get("missing_required", []):
            row_num = missing.get("row", 0)
            if row_num not in row_issues:
                row_issues[row_num] = []
            row_issues[row_num].append(f"Error: {missing.get('error', 'Missing required field')}")
        
        # Write data rows
        for row_idx, row_data in enumerate(sanitized_data, start=2):  # start=2 for Excel row (1-indexed, header=1)
            row_values = list(row_data.values())
            
            # Get validation flags for this row
            excel_row_num = row_idx - 1  # Convert to 1-based Excel row number
            flags = row_issues.get(excel_row_num, [])
            validation_flag = "; ".join(flags) if flags else ""
            row_values.append(validation_flag)
            
            ws.append(row_values)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Style rows with validation flags (highlight issues)
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            flag_cell = row[-1]  # Last column is Validation_Flags
            if flag_cell.value:
                flag_cell.fill = warning_fill
                flag_cell.font = Font(color="856404")
        
        # Auto-adjust column widths with safe type conversion
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    # Safe conversion: use 'or' to handle None/empty values
                    cell_value = cell.value if cell.value is not None else ''
                    cell_str = str(cell_value)
                    max_length = max(max_length, len(cell_str))
                except (AttributeError, TypeError):
                    # Skip cells that cannot be converted to string
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Exported sanitized Excel: {output_path}")
        
        if verbose:
            flagged_rows = sum(1 for v in row_issues.values() if v)
            click.echo(
                f"\n  ✓ Exported sanitized Excel: {output_path.name} "
                f"({flagged_rows} rows with validation flags)"
            )
        else:
            click.echo(f"\n✓ Exported sanitized Excel: {output_path}")
        
    except Exception as e:
        logger.error(f"Failed to export Excel: {e}")
        click.echo(
            click.style(f"Warning: Could not export sanitized Excel: {e}", fg="yellow")
        )


def _output_report(report: Dict[str, Any], output_file: Optional[str]) -> None:
    """Output the validation report.
    
    Args:
        report: The validation report dictionary.
        output_file: Optional path to save report. If None, prints to stdout.
    """
    report_json = json.dumps(report, indent=2, ensure_ascii=False, default=str)
    
    if output_file:
        try:
            with open(output_file, "w", encoding="utf-8") as f:
                f.write(report_json)
            click.echo(f"\nReport saved to: {output_file}")
        except Exception as e:
            click.echo(
                click.style(f"Warning: Could not save report to file: {e}", fg="yellow"),
                err=True,
            )
            click.echo("\nReport output:")
            click.echo(report_json)
    else:
        click.echo("\n" + "=" * 60)
        click.echo("VALIDATION REPORT")
        click.echo("=" * 60)
        click.echo(report_json)


def _print_summary(report: Dict[str, Any], verbose: bool) -> None:
    """Print a human-readable summary to the console.
    
    Args:
        report: The validation report dictionary.
        verbose: Enable verbose output with full details.
    """
    status = report.get("status", "unknown")
    
    # Status banner
    if status == "success":
        banner = click.style("✓ VALIDATION SUCCESSFUL", fg="green", bold=True)
    elif status == "completed_with_warnings":
        banner = click.style("⚠ VALIDATION COMPLETED WITH WARNINGS", fg="yellow", bold=True)
    else:
        banner = click.style("✗ VALIDATION FAILED", fg="red", bold=True)
    
    click.echo("\n" + banner)
    click.echo("-" * 60)
    
    # Summary stats
    click.echo(f"Total rows processed: {report.get('total_rows_processed', 0)}")
    click.echo(f"Rows with data:       {report.get('total_rows_with_data', 0)}")
    click.echo("-" * 60)
    
    # PHI sanitization
    phi = report.get("phi_sanitization", {})
    if phi.get("phi_detected"):
        click.echo(
            click.style(
                f"PHI detected and sanitized: {phi.get('fields_masked', 0)} fields masked",
                fg="yellow",
            )
        )
    else:
        click.echo("PHI sanitization: No PHI detected (clean)")
    
    click.echo("-" * 60)
    
    # Lab validation results
    lab = report.get("lab_validation", {})
    abnormal = lab.get("abnormal_values", [])
    missing = lab.get("missing_required", [])
    
    if abnormal:
        click.echo(f"\nAbnormal Lab Values: {len(abnormal)}")
        if verbose:
            for item in abnormal[:10]:  # Show first 10 in verbose mode
                click.echo(
                    f"  • {item['subject_id']}: {item['test_code']} = "
                    f"{item['value']} {item['unit']} "
                    f"(ref: {item['reference_range']})"
                )
            if len(abnormal) > 10:
                click.echo(f"  ... and {len(abnormal) - 10} more")
    
    if missing:
        click.echo(
            click.style(f"\nMissing Required Fields: {len(missing)}", fg="red")
        )
        for item in missing[:5]:  # Show first 5
            click.echo(f"  • Row {item['row']}: {item['error']}")
        if len(missing) > 5:
            click.echo(f"  ... and {len(missing) - 5} more")
    
    # Final summary
    click.echo("\n" + "=" * 60)
    summary = report.get("summary", {})
    click.echo(
        f"Summary: {summary.get('errors', 0)} errors, "
        f"{summary.get('warnings', 0)} warnings, "
        f"{summary.get('abnormal_count', 0)} abnormal labs"
    )
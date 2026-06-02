"""Data extractor module with EDC data extraction and PHI sanitization.

This module provides the EDCDataExtractor class for reading clinical trial data
from CSV and Excel files while automatically detecting and masking Protected
Health Information (PHI).
"""

import csv
import io
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

from medical_cli.utils.logger import setup_logger


# Configure module logger
logger = setup_logger("medical_cli.data_extractor")

# PHI detection patterns
PHI_PATTERNS: Dict[str, re.Pattern] = {
    # Chinese name pinyin initials (e.g., ZY, LLM, WangM)
    "name_pinyin": re.compile(
        r"\b([A-Z][a-z]{1,20}\s*){2,4}|[A-Z]{2,5}\s*[A-Z]{2,5}\b",
        re.IGNORECASE
    ),
    # Chinese ID number (18 digits)
    "id_number": re.compile(
        r"\b\d{15}(\d{2}[\dXx])?\b|\b\d{17}[\dXx]\b",
    ),
    # Phone numbers (various formats)
    "phone": re.compile(
        r"\b1[3-9]\d{9}\b|\b\(\d{3}\)\s*\d{3,4}[- ]?\d{4}\b|\b\d{3}[- ]\d{4}[- ]\d{4}\b",
    ),
    # Email addresses
    "email": re.compile(
        r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",
    ),
    # Medical record numbers (common patterns)
    "mrn": re.compile(
        r"\b(MRN|MR|REC|HIS)[#:=]?\d{5,}\b|\b\d{8,12}\b",
        re.IGNORECASE
    ),
    # Date of birth patterns that might identify patients
    "dob": re.compile(
        r"(DOB|Date\s*of\s*Birth|Birth)[=:]\s*\d{4}[-/]\d{1,2}[-/]\d{1,2}",
        re.IGNORECASE
    ),
    # Address patterns
    "address": re.compile(
        r"\d+\s+[\w\s]+(?:Street|St|Avenue|Ave|Road|Rd|Boulevard|Blvd|Lane|Ln|DR|Dr)\b",
        re.IGNORECASE
    ),
    # Social security numbers (US format)
    "ssn": re.compile(
        r"\b\d{3}[- ]\d{2}[- ]\d{4}\b",
    ),
    # Bank account patterns
    "bank_account": re.compile(
        r"\b\d{10,20}\b",
    ),
}

# Field names that commonly contain PHI
PHI_FIELD_NAMES: List[str] = [
    "patient_name", "name", "patient_id", "subject_name", "full_name",
    "first_name", "last_name", "given_name", "family_name", "surname",
    "birth_date", "dob", "birthday", "date_of_birth", "ssn", "social_security",
    "id_number", "id_card", "identity", "passport", "phone", "telephone",
    "mobile", "email", "address", "home_address", "mailing_address",
    "insurance", "insurance_number", "bank_account", "credit_card",
]

# Columns that should never contain PHI (safe columns)
SAFE_COLUMNS: List[str] = [
    "subject_id", "site", "visit_date", "visit", "age", "gender",
    "lab_value", "test_code", "result", "value", "unit", "reference_range",
]


class PHIDetectionResult:
    """Result of PHI detection scan.
    
    Attributes:
        detected: List of detected PHI items with location and type.
        masked_count: Number of values that were masked.
        is_clean: Whether no PHI was detected.
    """
    
    def __init__(self) -> None:
        """Initialize empty detection result."""
        self.detected: List[Dict[str, Any]] = []
        self.masked_count: int = 0
        self.is_clean: bool = True
    
    def add_detection(
        self,
        field: str,
        value: str,
        phi_type: str,
        row_index: int,
        masked_value: str,
    ) -> None:
        """Record a PHI detection.
        
        Args:
            field: Name of the field containing PHI.
            value: Original value that contained PHI.
            phi_type: Type of PHI detected.
            row_index: Row number where PHI was found.
            masked_value: Replacement value used.
        """
        self.detected.append({
            "field": field,
            "value": value,
            "type": phi_type,
            "row": row_index,
            "masked_to": masked_value,
        })
        self.masked_count += 1
        self.is_clean = False
        logger.debug(
            f"PHI detected: {phi_type} in field '{field}' "
            f"at row {row_index}"
        )
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert result to dictionary."""
        return {
            "detected_items": self.detected,
            "masked_count": self.masked_count,
            "is_clean": self.is_clean,
            "total_detections": len(self.detected),
        }


class EDCDataExtractor:
    """EDC (Electronic Data Capture) data extractor with PHI sanitization.
    
    This class reads clinical trial data from CSV and Excel files, performs
    automatic PHI detection and masking, and provides structured data access.
    
    Attributes:
        input_path: Path to the input file or directory.
        encoding: Character encoding for file reading (default: utf-8).
        strict_mode: If True, raises exceptions on PHI detection.
        detection_result: Result of the last PHI detection scan.
        
    Example:
        >>> extractor = EDCDataExtractor("./clinical_data.csv")
        >>> data = extractor.read_data()
        >>> print(f"Extracted {len(data)} records")
        >>> print(f"PHI detection: {extractor.detection_result.is_clean}")
    """
    
    MASK_VALUE: str = "*****"
    VERSION: str = "1.0.0"
    
    def __init__(
        self,
        input_path: str,
        encoding: str = "utf-8",
        strict_mode: bool = False,
    ) -> None:
        """Initialize the EDC data extractor.
        
        Args:
            input_path: Path to CSV or Excel file to read.
            encoding: Character encoding for file reading.
            strict_mode: If True, raise exception when PHI is detected.
            
        Raises:
            FileNotFoundError: If input file does not exist.
            ValueError: If file format is not supported.
        """
        self.input_path = Path(input_path)
        self.encoding = encoding
        self.strict_mode = strict_mode
        self.detection_result: Optional[PHIDetectionResult] = None
        self._raw_data: List[Dict[str, Any]] = []
        
        # Validate file exists and has supported format
        self._validate_input()
        
        logger.info(
            f"Initialized EDCDataExtractor for: {self.input_path} "
            f"(encoding={encoding}, strict={strict_mode})"
        )
    
    def _validate_input(self) -> None:
        """Validate input file exists and has supported format.
        
        Raises:
            FileNotFoundError: If file does not exist.
            ValueError: If file format is not supported.
        """
        if not self.input_path.exists():
            logger.error(f"Input file not found: {self.input_path}")
            raise FileNotFoundError(f"Input file not found: {self.input_path}")
        
        ext = self.input_path.suffix.lower()
        supported_formats = {".csv", ".xlsx", ".xls"}
        
        if ext not in supported_formats:
            error_msg = (
                f"Unsupported file format: {ext}. "
                f"Supported formats: {', '.join(supported_formats)}"
            )
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.debug(f"Input validation passed: {self.input_path}")
    
    def read_data(
        self,
        skip_empty_rows: bool = True,
        trim_whitespace: bool = True,
    ) -> List[Dict[str, Any]]:
        """Read and parse data from the input file.
        
        Automatically detects and masks PHI content during reading.
        
        Args:
            skip_empty_rows: Skip rows with no data.
            trim_whitespace: Trim whitespace from all values.
            
        Returns:
            List of dictionaries representing parsed data rows.
            
        Raises:
            Exception: If file cannot be read or parsed.
        """
        logger.info(f"Reading data from: {self.input_path}")
        self.detection_result = PHIDetectionResult()
        
        try:
            ext = self.input_path.suffix.lower()
            
            if ext == ".csv":
                self._raw_data = self._read_csv(
                    skip_empty=skip_empty_rows,
                    trim=trim_whitespace,
                )
            elif ext in (".xlsx", ".xls"):
                self._raw_data = self._read_excel(
                    skip_empty=skip_empty_rows,
                    trim=trim_whitespace,
                )
            
            # Perform PHI detection and masking
            self._sanitize_phi()
            
            logger.info(
                f"Successfully read {len(self._raw_data)} records "
                f"(PHI masked: {self.detection_result.masked_count})"
            )
            
            return self._raw_data
            
        except Exception as e:
            logger.error(f"Failed to read data: {e}")
            raise
    
    def _read_csv(
        self,
        skip_empty: bool = True,
        trim: bool = True,
    ) -> List[Dict[str, Any]]:
        """Read data from CSV file.
        
        Args:
            skip_empty: Skip empty rows.
            trim: Trim whitespace from values.
            
        Returns:
            List of dictionaries with parsed row data.
            
        Raises:
            Exception: If file cannot be read.
        """
        rows: List[Dict[str, Any]] = []
        
        try:
            with open(
                self.input_path,
                "r",
                encoding=self.encoding,
                newline="",
            ) as f:
                reader = csv.DictReader(f)
                
                for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is 1)
                    # Clean row data
                    cleaned_row: Dict[str, Any] = {}
                    has_data = False
                    
                    for key, value in row.items():
                        if value is not None:
                            if trim:
                                value = value.strip()
                            if value:
                                has_data = True
                            cleaned_row[key] = value
                        else:
                            cleaned_row[key] = ""
                    
                    if has_data or not skip_empty:
                        rows.append(cleaned_row)
                        logger.debug(f"Read CSV row {row_num}")
                
            logger.info(f"CSV read complete: {len(rows)} rows")
            return rows
            
        except UnicodeDecodeError as e:
            logger.warning(
                f"Encoding error with {self.encoding}, trying 'latin-1': {e}"
            )
            # Fallback to latin-1 encoding
            with open(
                self.input_path,
                "r",
                encoding="latin-1",
                newline="",
            ) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cleaned_row = {
                        k: v.strip() if trim and v else v 
                        for k, v in row.items()
                    }
                    if any(cleaned_row.values()) or not skip_empty:
                        rows.append(cleaned_row)
            return rows
            
        except Exception as e:
            logger.error(f"CSV read error: {e}")
            raise
    
    def _read_excel(
        self,
        skip_empty: bool = True,
        trim: bool = True,
    ) -> List[Dict[str, Any]]:
        """Read data from Excel file (.xlsx/.xls).
        
        Args:
            skip_empty: Skip empty rows.
            trim: Trim whitespace from values.
            
        Returns:
            List of dictionaries with parsed row data.
            
        Raises:
            Exception: If file cannot be read.
        """
        rows: List[Dict[str, Any]] = []
        
        try:
            # Try to import openpyxl
            try:
                from openpyxl import load_workbook
            except ImportError:
                logger.error(
                    "openpyxl not installed. Install with: pip install openpyxl"
                )
                raise ImportError(
                    "Excel support requires openpyxl. "
                    "Install with: pip install openpyxl"
                )
            
            wb = load_workbook(filename=str(self.input_path), data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers: List[str] = []
            for cell in ws[1]:
                value = cell.value
                if value is not None:
                    headers.append(str(value).strip() if trim else str(value))
                else:
                    headers.append(f"Column_{cell.column}")
            
            logger.debug(f"Excel headers: {headers}")
            
            # Read data rows
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                cleaned_row: Dict[str, Any] = {}
                has_data = False
                
                for header, cell in zip(headers, row):
                    value = cell.value
                    
                    # Convert datetime objects to string
                    if isinstance(value, (datetime, date)):
                        value = value.strftime("%Y-%m-%d")
                    
                    if value is not None:
                        if trim and isinstance(value, str):
                            value = value.strip()
                        if value:
                            has_data = True
                        cleaned_row[header] = value
                    else:
                        cleaned_row[header] = ""
                
                if has_data or not skip_empty:
                    rows.append(cleaned_row)
                    logger.debug(f"Read Excel row {row_num}")
            
            wb.close()
            logger.info(f"Excel read complete: {len(rows)} rows")
            return rows
            
        except Exception as e:
            logger.error(f"Excel read error: {e}")
            raise
    
    def _sanitize_phi(self) -> None:
        """Sanitize PHI from the loaded data.
        
        Scans all string values for PHI patterns and replaces detected
        PHI with mask characters.
        """
        logger.info("Starting PHI sanitization...")
        sanitized_count = 0
        
        for row_idx, row in enumerate(self._raw_data):
            for field, value in row.items():
                if not isinstance(value, str):
                    continue
                
                if self._is_safe_column(field):
                    continue
                
                # Check for PHI in field name
                if self._is_phi_field_name(field):
                    row[field] = self.MASK_VALUE
                    self.detection_result.add_detection(
                        field=field,
                        value=value,
                        phi_type="phi_field_name",
                        row_index=row_idx,
                        masked_value=self.MASK_VALUE,
                    )
                    sanitized_count += 1
                    continue
                
                # Check for PHI in value
                masked_value, phi_info = self._detect_phi_in_value(value, field)
                
                if masked_value != value:
                    row[field] = masked_value
                    self.detection_result.add_detection(
                        field=field,
                        value=value,
                        phi_type=phi_info,
                        row_index=row_idx,
                        masked_value=masked_value,
                    )
                    sanitized_count += 1
        
        logger.info(
            f"PHI sanitization complete: {sanitized_count} values masked"
        )
        
        if self.strict_mode and not self.detection_result.is_clean:
            raise ValueError(
                f"PHI detected in strict mode. "
                f"Found {self.detection_result.masked_count} PHI items."
            )
    
    def _is_safe_column(self, field: str) -> bool:
        """Check if column is known to be safe (no PHI).
        
        Args:
            field: Column name to check.
            
        Returns:
            True if column is safe, False otherwise.
        """
        field_lower = field.lower()
        return any(
            safe in field_lower 
            for safe in SAFE_COLUMNS
        )
    
    def _is_phi_field_name(self, field: str) -> bool:
        """Check if field name commonly contains PHI.
        
        Args:
            field: Field name to check.
            
        Returns:
            True if field is likely to contain PHI.
        """
        field_lower = field.lower().replace(" ", "_").replace("-", "_")
        
        for phi_pattern in PHI_FIELD_NAMES:
            if phi_pattern in field_lower:
                logger.debug(f"PHI field name detected: {field}")
                return True
        
        return False
    
    def _detect_phi_in_value(
        self,
        value: str,
        field: str,
    ) -> tuple[str, str]:
        """Detect PHI in a string value.
        
        Args:
            value: String value to scan.
            field: Field name containing the value.
            
        Returns:
            Tuple of (masked_value, phi_type) or (original_value, "") if no PHI.
        """
        if not value or len(value) < 2:
            return value, ""
        
        # Check against each PHI pattern
        for phi_type, pattern in PHI_PATTERNS.items():
            matches = pattern.findall(value)
            if matches:
                # Replace the matched portion with mask
                masked = pattern.sub(self.MASK_VALUE, value)
                return masked, phi_type
        
        return value, ""
    
    def get_phi_report(self) -> Dict[str, Any]:
        """Get a report of PHI detection results.
        
        Returns:
            Dictionary containing PHI detection statistics and details.
        """
        if self.detection_result is None:
            return {
                "scanned": False,
                "message": "No data has been scanned yet. Call read_data() first.",
            }
        
        report = self.detection_result.to_dict()
        report["input_file"] = str(self.input_path)
        report["total_records"] = len(self._raw_data)
        
        # Group detections by type
        by_type: Dict[str, int] = {}
        for item in self.detection_result.detected:
            phi_type = item["type"]
            by_type[phi_type] = by_type.get(phi_type, 0) + 1
        
        report["detections_by_type"] = by_type
        
        return report
    
    def extract_subset(
        self,
        fields: List[str],
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """Extract a subset of columns and rows.
        
        Args:
            fields: List of field names to extract.
            start_row: Starting row index (0-based), None for beginning.
            end_row: Ending row index (exclusive), None for end.
            
        Returns:
            List of dictionaries with only the selected fields.
        """
        if not self._raw_data:
            logger.warning("No data loaded. Call read_data() first.")
            return []
        
        subset: List[Dict[str, Any]] = []
        
        for row in self._raw_data[start_row:end_row]:
            subset_row = {field: row.get(field, "") for field in fields}
            subset.append(subset_row)
        
        logger.debug(
            f"Extracted subset: {len(fields)} fields, "
            f"{len(subset)} rows"
        )
        
        return subset
    
    def to_json(self) -> str:
        """Convert loaded data to JSON string.
        
        Returns:
            JSON string representation of the data.
        """
        import json
        
        if not self._raw_data:
            return "[]"
        
        return json.dumps(self._raw_data, indent=2, ensure_ascii=False)
    
    def to_csv(self, output_path: Optional[str] = None) -> str:
        """Convert loaded data to CSV format.
        
        Args:
            output_path: Optional path to write CSV file.
            
        Returns:
            CSV string representation of the data.
        """
        if not self._raw_data:
            return ""
        
        output = io.StringIO() if output_path is None else open(
            output_path, "w", encoding=self.encoding, newline=""
        )
        
        writer = csv.DictWriter(output, fieldnames=self._raw_data[0].keys())
        writer.writeheader()
        writer.writerows(self._raw_data)
        
        if output_path is None:
            return output.getvalue()
        
        output.close()
        return output_path


# Module-level utility functions for backwards compatibility
def extract(input_path: str, output_path: str, fmt: str = "json") -> int:
    """Extract clinical data from source files.
    
    Args:
        input_path: Path to input file or directory.
        output_path: Path to output file.
        fmt: Output format (json, csv, xml).
        
    Returns:
        Exit code (0 for success, non-zero for failure).
    """
    try:
        extractor = EDCDataExtractor(input_path)
        data = extractor.read_data()
        
        if fmt == "json":
            import json
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        elif fmt == "csv":
            import csv
            if data:
                with open(output_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
        else:
            logger.error(f"Unsupported format: {fmt}")
            return 1
        
        logger.info(f"Data extracted to: {output_path}")
        return 0
        
    except Exception as e:
        logger.error(f"Extraction failed: {e}")
        return 1


def extract_from_file(file_path: str) -> Dict[str, Any]:
    """Extract data from a single file.
    
    Args:
        file_path: Path to the source file.
        
    Returns:
        Extracted data as dictionary with 'records' and 'metadata' keys.
    """
    try:
        extractor = EDCDataExtractor(file_path)
        records = extractor.read_data()
        
        return {
            "records": records,
            "metadata": {
                "file": file_path,
                "record_count": len(records),
                "phi_clean": extractor.detection_result.is_clean,
            },
        }
        
    except Exception as e:
        logger.error(f"File extraction failed: {e}")
        return {"records": [], "metadata": {"error": str(e)}}


def extract_from_directory(dir_path: str) -> List[Dict[str, Any]]:
    """Extract data from all files in a directory.
    
    Args:
        dir_path: Path to the source directory.
        
    Returns:
        List of all extracted records from all files.
    """
    all_records: List[Dict[str, Any]] = []
    
    path = Path(dir_path)
    if not path.is_dir():
        logger.error(f"Not a directory: {dir_path}")
        return all_records
    
    # Find all CSV and Excel files
    for ext in [".csv", ".xlsx", ".xls"]:
        for file_path in path.glob(f"*{ext}"):
            try:
                logger.info(f"Processing: {file_path}")
                extractor = EDCDataExtractor(str(file_path))
                records = extractor.read_data()
                
                # Add source file info to each record
                for record in records:
                    record["_source_file"] = str(file_path.name)
                
                all_records.extend(records)
                
            except Exception as e:
                logger.warning(f"Failed to process {file_path}: {e}")
                continue
    
    logger.info(f"Directory extraction complete: {len(all_records)} total records")
    return all_records